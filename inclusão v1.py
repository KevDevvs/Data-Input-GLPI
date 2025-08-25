import requests
from openpyxl import load_workbook

# ==== CONFIGURAÇÃO ====
GLPI_URL = "http://localhost/glpi/apirest.php"  # URL da API do GLPI
API_USER_TOKEN = "6I5gkjwvKcVaOpQOw9TqUKaX0119owQuOmJzENHd"  # Token do usuário gerado no GLPI
API_APP_TOKEN = "xOTvbn5CqY1tfQXrHdSPmUgHSBd9y6mIZyUrt9Z8" # Token da aplicação gerado no GLPI
XLSX_PATH = "dados_glpi.xlsx"        # Caminho da planilha de entrada

# ==== AUTENTICAÇÃO ====
def init_session():
    """Inicia sessão no GLPI e retorna o token de sessão"""
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"user_token {API_USER_TOKEN}",
        "App-Token": API_APP_TOKEN
    }
    print("[INICIANDO] Autenticando no GLPI...")
    resp = requests.get(f"{GLPI_URL}/initSession", headers=headers)
    resp.raise_for_status()
    print("[OK] Sessão iniciada!")
    return resp.json()["session_token"]

def kill_session(session_token):
    """Encerra a sessão no GLPI"""
    headers = {"Session-Token": session_token, "App-Token": API_APP_TOKEN}
    print("[FINALIZANDO] Encerrando sessão no GLPI...")
    requests.get(f"{GLPI_URL}/killSession", headers=headers)
    print("[OK] Sessão encerrada!")

# ==== CACHE ====
entity_cache = {}
group_cache = {}
user_cache = {}

# ==== FUNÇÕES AUXILIARES ====
def get_or_create_entity(session_token, name, parent_id=0):
    """
    Busca ou cria uma entidade no GLPI.
    Usa cache para evitar buscas repetidas.
    """
    if not name:
        return parent_id
    key = (name, parent_id)
    if key in entity_cache:
        print(f"[CACHE] Entidade '{name}' já existe (ID: {entity_cache[key]})")
        return entity_cache[key]

    headers = {"Session-Token": session_token, "App-Token": API_APP_TOKEN}
    print(f"[BUSCA] Procurando entidade '{name}'...")
    search = requests.get(
        f"{GLPI_URL}/search/Entity",
        headers=headers,
        params={"criteria[0][field]": 1, "criteria[0][searchtype]": "equals", "criteria[0][value]": name}
    ).json()

    if search.get("totalcount", 0) > 0:
        entity = search["data"][0]
        entity_id = int(entity.get("id", entity.get("2", 0)))
        print(f"[OK] Entidade '{name}' encontrada (ID: {entity_id})")
        # Se estiver deletada, restaurar
        if entity.get("is_deleted", False):
            print(f"[RESTAURO] Restaurando entidade '{name}'...")
            requests.put(f"{GLPI_URL}/Entity/{entity_id}", json={"input": {"is_deleted": 0}}, headers=headers)
        # Atualizar entidade
        requests.put(f"{GLPI_URL}/Entity/{entity_id}", json={"input": {"name": name, "entities_id": parent_id}}, headers=headers)
    else:
        print(f"[CRIANDO] Entidade '{name}'...")
        data = {"input": {"name": name, "entities_id": parent_id}}
        created = requests.post(f"{GLPI_URL}/Entity", json=data, headers=headers).json()
        entity_id = created["id"]
        print(f"[OK] Entidade '{name}' criada (ID: {entity_id})")

    entity_cache[key] = entity_id
    return entity_id

def get_or_create_group(session_token, name, entity_id):
    """Busca ou cria um grupo no GLPI"""
    if not name:
        return None
    key = (name, entity_id)

    headers = {"Session-Token": session_token, "App-Token": API_APP_TOKEN}
    print(f"[BUSCA] Procurando grupo '{name}'...")
    search = requests.get(
        f"{GLPI_URL}/search/Group",
        headers=headers,
        params={"criteria[0][field]": 1, "criteria[0][searchtype]": "equals", "criteria[0][value]": name}
    ).json()

    if search.get("totalcount", 0) > 0:
        group = search["data"][0]
        group_id = int(group.get("id", group.get("2", 0)))
        print(f"[OK] Grupo '{name}' encontrado (ID: {group_id})")
        # Se estiver deletado, restaurar
        if group.get("is_deleted", False):
            print(f"[RESTAURO] Restaurando grupo '{name}'...")
            requests.put(f"{GLPI_URL}/Group/{group_id}", json={"input": {"is_deleted": 0}}, headers=headers)
        # Atualizar grupo
        requests.put(f"{GLPI_URL}/Group/{group_id}", json={"input": {"name": name, "entities_id": entity_id}}, headers=headers)
    else:
        print(f"[CRIANDO] Grupo '{name}'...")
        data = {"input": {"name": name, "entities_id": entity_id}}
        created = requests.post(f"{GLPI_URL}/Group", json=data, headers=headers).json()
        group_id = created["id"]
        print(f"[OK] Grupo '{name}' criado (ID: {group_id})")

    group_cache[key] = group_id
    return group_id

def get_or_create_user(session_token, name, profile_id, group_id, entity_id):
    """Busca ou cria um usuário no GLPI e associa a grupo e perfil"""
    if not name or not profile_id:
        return None
    key = (name, profile_id, group_id, entity_id)
    if key in user_cache:
        print(f"[CACHE] Usuário '{name}' já existe (ID: {user_cache[key]})")
        return user_cache[key]

    headers = {"Session-Token": session_token, "App-Token": API_APP_TOKEN}
    print(f"[BUSCA] Procurando usuário '{name}'...")
    search = requests.get(
        f"{GLPI_URL}/search/User",
        headers=headers,
        params={"criteria[0][field]": 1, "criteria[0][searchtype]": "equals", "criteria[0][value]": name}
    ).json()

    if search.get("totalcount", 0) > 0:
        user = search["data"][0]
        user_id = int(user.get("id", user.get("2", 0)))
        print(f"[OK] Usuário '{name}' encontrado (ID: {user_id})")
        # Se estiver deletado, restaurar
        if user.get("is_deleted", False):
            print(f"[RESTAURO] Restaurando usuário '{name}'...")
            requests.put(f"{GLPI_URL}/User/{user_id}", json={"input": {"is_deleted": 0}}, headers=headers)
        # Atualizar usuário
        update_data = {
            "input": {
                "name": name,
                "realname": name,
                "entities_id": entity_id,
                "_profiles_id": profile_id,
                "groups_id": group_id
            }
        }
        requests.put(f"{GLPI_URL}/User/{user_id}", json=update_data, headers=headers)
    else:
        print(f"[CRIANDO] Usuário '{name}'...")
        data = {
            "input": {
                "name": name,
                "realname": name,
                "entities_id": entity_id,
                "_profiles_id": profile_id,
                "groups_id": group_id
            }
        }
        created = requests.post(f"{GLPI_URL}/User", json=data, headers=headers).json()
        if "id" in created:
            user_id = created["id"]
            print(f"[OK] Usuário '{name}' criado (ID: {user_id})")
        else:
            print(f"[ERRO] Não foi possível criar o usuário '{name}'. Resposta da API:")
            print(created)
            user_id = None

    user_cache[key] = user_id
    return user_id

def create_asset(session_token, asset_type, data):
    """Cria um ativo no GLPI (Computer, Phone, Line, etc.)"""
    headers = {"Session-Token": session_token, "App-Token": API_APP_TOKEN}
    print(f"[BUSCA] Procurando {asset_type} '{data.get('name', '')}'...")
    # Busca por nome e entidade
    search = requests.get(
        f"{GLPI_URL}/search/{asset_type}",
        headers=headers,
        params={"criteria[0][field]": 1, "criteria[0][searchtype]": "equals", "criteria[0][value]": data.get("name", ""), "criteria[1][field]": 4, "criteria[1][searchtype]": "equals", "criteria[1][value]": data.get("entities_id", 0)}
    ).json()

    if search.get("totalcount", 0) > 0:
        asset = search["data"][0]
        asset_id = int(asset.get("id", asset.get("2", 0)))
        print(f"[OK] {asset_type} '{data.get('name', '')}' encontrado (ID: {asset_id})")
        # Se estiver deletado, restaurar
        if asset.get("is_deleted", False):
            print(f"[RESTAURO] Restaurando {asset_type} '{data.get('name', '')}'...")
            requests.put(f"{GLPI_URL}/{asset_type}/{asset_id}", json={"input": {"is_deleted": 0}}, headers=headers)
        # Atualizar ativo
        requests.put(f"{GLPI_URL}/{asset_type}/{asset_id}", json={"input": data}, headers=headers)
    else:
        print(f"[CRIANDO] {asset_type} '{data.get('name', '')}'...")
        resp = requests.post(f"{GLPI_URL}/{asset_type}", json={"input": data}, headers=headers).json()
        asset_id = resp.get("id")
        print(f"[OK] {asset_type} '{data.get('name', '')}' criado (ID: {asset_id})")
    return asset_id

# ==== PROCESSAMENTO ====
def process_excel():
    """Lê a planilha XLSX e cria entidades, grupos, usuários e ativos no GLPI"""
    print("[INICIANDO] Processamento da planilha XLSX...")
    session_token = init_session()
    print(f"[OK] Lendo arquivo '{XLSX_PATH}'...")
    wb = load_workbook(XLSX_PATH)
    sheet = wb.active

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        print(f"\n[LINHA {idx}] Processando dados...")
        (
            nome_user, perfil, perfil_id, grupo,
            entidade_a, entidade_b, entidade_c, linha,
            modelo_cel, imei,
            modelo_nb, marca_nb, serial_nb
        ) = row

        # Criação da hierarquia de entidades (cascata)
        entidade_a_id = get_or_create_entity(session_token, entidade_a)
        entidade_b_id = get_or_create_entity(session_token, entidade_b, entidade_a_id) if entidade_b else None
        entidade_c_id = get_or_create_entity(session_token, entidade_c, entidade_b_id) if entidade_c else None

        # A entidade final será a mais profunda criada
        final_entity_id = entidade_c_id or entidade_b_id or entidade_a_id

        # Criação do grupo
        grupo_id = get_or_create_group(session_token, grupo, final_entity_id)

        # Criação do usuário (se informado)
        user_id = get_or_create_user(session_token, nome_user, perfil_id, grupo_id, final_entity_id)

        # Criar Linha como ativo do tipo "Line"
        if linha:
            create_asset(session_token, "Line", {
                "name": linha,
                "entities_id": final_entity_id,
                "users_id": user_id if user_id else 0
            })

        # Criar Celular como ativo do tipo "Phone"
        if modelo_cel:
            create_asset(session_token, "Phone", {
                "name": modelo_cel,
                "serial": imei,
                "entities_id": final_entity_id,
                "users_id": user_id if user_id else 0
            })

        # Criar Notebook como ativo do tipo "Computer"
        if modelo_nb:
            create_asset(session_token, "Computer", {
                "name": modelo_nb,
                "serial": serial_nb,
                "brand": marca_nb,
                "entities_id": final_entity_id,
                "users_id": user_id if user_id else 0
            })

        print(f"[OK] Linha {idx} processada!")

    print("\n[FINALIZANDO] Todas as linhas processadas!")
    kill_session(session_token)
    print("[OK] Processo concluído!")

if __name__ == "__main__":
    import traceback
    try:
        process_excel()
    except Exception as e:
        print("\n[ERRO] Ocorreu uma exceção durante o processamento!")
        traceback.print_exc()
