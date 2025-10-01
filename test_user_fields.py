"""
Script de teste para validar as melhorias nos campos de usuário
"""
import openpyxl
from helper.colors import c

def test_new_user_fields():
    """Testa se os novos campos de usuário estão sendo lidos corretamente"""
    print(c("🧪 Testando novos campos de usuário...", 'yellow'))
    
    try:
        # Carrega a planilha
        wb = openpyxl.load_workbook("input_glpi.xlsx")
        sheet = wb.active
        
        print(c("\n📊 Teste dos novos campos de usuário:", 'cyan'))
        print(c("Linha | Nome | Email | Email Corp | Cel. Pessoal | Posição | Comentário", 'white'))
        print(c("-" * 80, 'white'))
        
        # Testa as primeiras 5 linhas com dados
        for idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
            if len(row) >= 7:  # Verifica se tem colunas suficientes
                nome = row[0] if len(row) > 0 else None
                email = row[1] if len(row) > 1 else None 
                cpf = row[2] if len(row) > 2 else None
                email_corp = row[3] if len(row) > 3 else None
                celular_pessoal = row[4] if len(row) > 4 else None
                posicao = row[5] if len(row) > 5 else None
                comentario = row[6] if len(row) > 6 else None
                
                # Converte para string para exibição
                nome_str = str(nome)[:20] if nome else "None"
                email_str = str(email)[:15] if email else "None"
                email_corp_str = str(email_corp)[:15] if email_corp else "None"
                cel_str = str(celular_pessoal)[:12] if celular_pessoal else "None"
                pos_str = str(posicao)[:15] if posicao else "None"
                com_str = str(comentario)[:15] if comentario else "None"
                
                print(f"{idx:4d} | {nome_str:20} | {email_str:15} | {email_corp_str:15} | {cel_str:12} | {pos_str:15} | {com_str:15}")
                
                # Testa a lógica de priorização do email
                email_principal = email_corp if email_corp and str(email_corp).strip() else email
                if email_principal:
                    print(f"      🎯 Email principal selecionado: {email_principal}")
                print()
                
    except Exception as e:
        print(c(f"❌ Erro ao testar novos campos: {e}", 'red'))

def test_structure_integrity():
    """Testa se a estrutura da planilha está íntegra"""
    print(c("🔍 Verificando integridade da estrutura...", 'yellow'))
    
    try:
        wb = openpyxl.load_workbook("input_glpi.xlsx")
        sheet = wb.active
        
        expected_columns = [
            "Nome do User", "email", "CPF", "Email corp", "Celular pessoal", 
            "Posição", "Comentario", "Status", "Raiz", "new BU", "cliente", 
            "Projeto", "Dimensoes", "Linha", "Operadora"
        ]
        
        first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        print(c("✅ Verificação das primeiras 15 colunas:", 'green'))
        for i, expected in enumerate(expected_columns):
            actual = first_row[i] if i < len(first_row) else None
            status = "✅" if actual == expected else "❌"
            print(f"  {status} Coluna {i+1:2d}: esperado '{expected}', encontrado '{actual}'")
            
        print(c(f"\n📊 Total de colunas: {len([h for h in first_row if h])}", 'cyan'))
        print(c("✅ Verificação concluída!", 'green'))
        
    except Exception as e:
        print(c(f"❌ Erro na verificação: {e}", 'red'))

if __name__ == "__main__":
    test_new_user_fields()
    print("\n" + "="*80 + "\n")
    test_structure_integrity()