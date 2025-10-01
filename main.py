
import requests
from create_info.glpi_objects.component import link_component
from remove_data.remove_data import reset_glpi
from helper.colors import c
from create_info.create_entity_hierarchy import create_entity_hierarchy
from create_info.create_users import create_user
from glpi_session.glpi_session import init_session, kill_session
from create_info.create_asset import create_asset
from create_info.get_or_create import get_or_create_manufacturer, get_or_create_model, get_or_create
from create_info.management import get_or_create_contract, link_contract_to_asset, create_management_info, get_or_create_supplier
from helper.read_config import GLPI_URL, HEADERS
from helper.logger import get_logger, close_logger
import openpyxl
import openpyxl
import os
from helper.read_config import FILE_PATH

total_processado = 0
total_sucesso = 0
total_erro = 0
session = None

def update_status_column(wb, row_idx, column_idx, status, description=None):
    """
    Atualiza uma coluna de status na planilha
    
    Args:
        wb: Workbook do openpyxl
        row_idx: Índice da linha (1-based)
        column_idx: Índice da coluna (1-based) 
        status: Status básico ("OK", "ERRO", "" ou descrição específica)
        description: Descrição específica do erro (opcional)
    """
    try:
        sheet = wb.active
        
        # Se tem descrição específica, usa ela; senão usa o status
        final_status = description if description else status
        
        sheet.cell(row=row_idx, column=column_idx).value = final_status
        return True
    except Exception as e:
        logger = get_logger()
        logger.error(f"Erro ao atualizar status na coluna {column_idx}: {e}")
        return False

def save_excel_file(wb):
    """Salva o arquivo Excel com as atualizações de status"""
    try:
        wb.save(FILE_PATH)
        return True
    except Exception as e:
        print(c(f"❌ Erro ao salvar planilha: {e}", 'red'))
        return False

def main():
    global total_processado, total_sucesso, total_erro, session
    
    # Inicializa o logger
    logger = get_logger()
    
    ### Valida se o arquivo existe
    try:
        logger.process_start("Processo de Input de dados no GLPI")

        if not os.path.exists(FILE_PATH):
            logger.error(f"Arquivo '{FILE_PATH}' não encontrado!")
            return total_processado, total_sucesso, total_erro
            
        wb = openpyxl.load_workbook(FILE_PATH)
        sheet = wb.active
        if sheet.max_row < 2:
            logger.error("Planilha vazia ou sem dados!")
            return
        session = init_session()
    except Exception as e:
        logger.error(f"Erro ao processar arquivo: {str(e)}")
        if session:
            kill_session(session)
        return

    ### Itera sobre as linhas da planilha, pulando o cabeçalho
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        logger.line_processing(idx)
        try:
            # Desempacota os campos da linha (incluindo novos campos de usuário)
            nome, email, cpf, email_corp, celular_pessoal, posicao, comentario_user, status_user, ent_a, ent_b, ent_c, ent_d, ent_comment, linha, linha_operadora, contrato_linha, line_status, line_type, fornecedor_linha, data_inicial_linha, valor_linha, cel_type, cel_marca, cel_modelo, cel_imei, cel_status, cel_coment, nb_marca, nb_modelo, nb_type, nb_serial, nb_ativo, nb_armazenamento, nb_processador, nb_memoria, contrato_notebook, comprado_em_notebook, fornecedor_notebook, nb_coment, nb_status, input_user, input_line, input_mobile, input_notebook = row

            
            # Cria entidades em cascata e pega o ID do último nível preenchido
            if not ent_a:
                logger.error(f"Entidade A obrigatória na linha {idx}")
                continue
            
            entidade_final_id = None
            entidade_final_id = create_entity_hierarchy(session, ent_a, ent_b, ent_c, ent_d, ent_comment)

            if not entidade_final_id:
                logger.error(f"Falha ao criar hierarquia de entidades na linha {idx}")
                continue

            # Cria usuário e vincula sempre ao grupo 'User' e ao perfil Self-Service
            user_id = None
            if nome and str(nome).strip():
                # Prioriza email corporativo se disponível, senão usa email padrão
                email_principal = email_corp if email_corp and str(email_corp).strip() else email
                
                # Validação de email obrigatório
                if not email_principal or not email_principal.strip():
                    error_msg = "Email obrigatório ausente"
                    logger.error(f"{error_msg} para o usuário '{nome}' na linha {idx}")
                    update_status_column(wb, idx, 41, "ERRO", error_msg)
                    continue
                
                # Verifica se o email tem formato válido (deve conter @)
                email_clean = str(email_principal).strip()
                if '@' not in email_clean:
                    error_msg = "Email inválido (sem @)"
                    logger.error(f"{error_msg}: '{email_clean}' para o usuário '{nome}' na linha {idx}")
                    update_status_column(wb, idx, 41, "ERRO", error_msg)
                    continue
                
                perfil_id = 1  # ID do perfil a ser vinculado
                
                # Prepara email para uso (adiciona @ se necessário)
                email_param = f"@{email_clean}" if not email_clean.startswith('@') else email_clean
                
                # Tratamento de CPF (agora opcional)
                cpf_formatado = None
                if cpf and str(cpf).strip():
                    cpf_formatado = str(cpf).zfill(11)
                    logger.debug(f"CPF fornecido e formatado: {cpf_formatado}")
                else:
                    logger.debug(f"CPF não fornecido - campo opcional")
                
                # Processa informações adicionais do usuário
                celular_formatado = None
                if celular_pessoal and str(celular_pessoal).strip():
                    celular_formatado = str(celular_pessoal).strip()
                    logger.debug(f"Celular pessoal fornecido: {celular_formatado}")
                
                posicao_formatada = None
                if posicao and str(posicao).strip():
                    posicao_formatada = str(posicao).strip()
                    logger.debug(f"Posição fornecida: {posicao_formatada}")
                    
                comentario_formatado = None
                if comentario_user and str(comentario_user).strip():
                    comentario_formatado = str(comentario_user).strip()
                    logger.debug(f"Comentário fornecido: {comentario_formatado}")

                user_id, user_error = create_user(
                    session, nome, email_param, perfil_id, entidade_final_id, 
                    status_user, cpf_formatado, celular_formatado, 
                    posicao_formatada, comentario_formatado
                )
                
                # Atualiza status do usuário na planilha
                if user_id:
                    update_status_column(wb, idx, 41, "OK")  # Coluna 41 = Input User
                    logger.success(f"Usuário criado (ID: {user_id}) - Status atualizado: OK")
                else:
                    error_msg = user_error if user_error else "Falha ao criar usuário"
                    update_status_column(wb, idx, 41, "ERRO", error_msg)
                    logger.error(f"Falha ao criar usuário: {error_msg} - Status atualizado: {error_msg}")
            else:
                # Se não há usuário para processar, marca como vazio
                if not nome or str(nome).strip() == "":
                    update_status_column(wb, idx, 41, "")
                    logger.debug(f"Sem usuário para processar na linha {idx}")

            # Cria ativos vinculados à entidade/usuário (apenas se campo preenchido)
            if linha:
                # Validação: verifica se operadora foi informada
                if not linha_operadora or not str(linha_operadora).strip():
                    error_msg = "Operadora não informada"
                    logger.error(f"Linha '{linha}' não pode ser criada: {error_msg}")
                    update_status_column(wb, idx, 42, "ERRO", error_msg)
                else:
                    # Busca ou cria a operadora
                    operators_response = requests.get(f"{GLPI_URL}/LineOperator", headers={**HEADERS, "Session-Token": session})
                    operators_list = operators_response.json()
                    
                    operator_id = None
                    
                    # Procura na lista de operadoras
                    if isinstance(operators_list, list):
                        for operator in operators_list:
                            if operator.get("name") == str(linha_operadora).strip():
                                operator_id = operator.get("id")
                                break
                                    
                    if operator_id and operator_id > 0:  # Garante que o ID é válido
                        print(c(f"✅ [OK] Operadora '{linha_operadora}' vinculada com sucesso", 'green'))

                        line_data = {
                        "name": linha,
                        "entities_id": entidade_final_id,
                        "users_id": user_id if user_id else 0,  # Usa 0 como fallback
                        "lineoperators_id": operator_id,
                        "linetypes_id": line_type,
                        "states_id": line_status  # Para linhas, o campo correto é states_id
                        }
                        
                        # Adiciona data inicial se fornecida
                        if data_inicial_linha and str(data_inicial_linha).strip():
                            # Converte data do formato DD/MM/YYYY para YYYY-MM-DD se necessário
                            data_formatada = str(data_inicial_linha).strip()
                            if '/' in data_formatada:
                                try:
                                    from datetime import datetime
                                    data_obj = datetime.strptime(data_formatada, '%d/%m/%Y')
                                    data_formatada = data_obj.strftime('%Y-%m-%d')
                                except:
                                    print(c(f"⚠️ Formato de data inválido: {data_formatada}", 'yellow'))
                            line_data["buy_date"] = data_formatada
                            print(c(f"📅 Data inicial adicionada: {data_formatada}", 'cyan'))

                        # Cria a linha primeiro
                        line_id, line_error = create_asset(session, "Line", line_data)
                        
                        # Atualiza status da linha na planilha
                        if line_id:
                            update_status_column(wb, idx, 42, "OK")  # Coluna 42 = Input Line
                            logger.item_created("Line", line_id, linha)
                        else:
                            error_msg = line_error if line_error else "Falha ao criar linha"
                            update_status_column(wb, idx, 42, "ERRO", error_msg)
                            logger.item_failed("Line", error_msg)
                        
                        # Se a linha foi criada com sucesso, processa informações adicionais
                        if line_id:
                            # Processa contrato da linha
                            if contrato_linha and str(contrato_linha).strip():
                                supplier_id = None
                                
                                # Primeiro cria/busca o fornecedor se fornecido (na entidade raiz)
                                if fornecedor_linha and str(fornecedor_linha).strip():
                                    supplier_id = get_or_create_supplier(session, str(fornecedor_linha).strip())
                                
                                # Cria/busca o contrato com o fornecedor (na entidade raiz)
                                contract_id = get_or_create_contract(session, str(contrato_linha).strip(), supplier_id=supplier_id)
                                if contract_id:
                                    link_contract_to_asset(session, "Line", line_id, contract_id)
                            
                            # Adiciona informações de Management
                            if data_inicial_linha or valor_linha:
                                create_management_info(
                                    session, 
                                    "Line", 
                                    line_id,
                                    buy_date=data_inicial_linha if data_inicial_linha and str(data_inicial_linha).strip() else None,
                                    value=valor_linha if valor_linha and str(valor_linha).strip() else None,
                                    supplier_id=supplier_id
                                )
                            
                            # Processa fornecedor da linha (caso não tenha contrato mas tenha fornecedor)
                            if not contrato_linha and fornecedor_linha and str(fornecedor_linha).strip():
                                supplier_id = get_or_create_supplier(session, str(fornecedor_linha).strip())
                                # Fornecedor criado na entidade raiz para uso futuro ou outros propósitos
                    else:
                        error_msg = f"Operadora '{linha_operadora}' não encontrada"
                        logger.error(f"Não foi possível encontrar a operadora '{linha_operadora}'")
                        update_status_column(wb, idx, 42, "ERRO", error_msg)  # Erro na operadora = erro na linha
            else:
                # Se não há linha para processar, marca como vazio
                if not linha or str(linha).strip() == "":
                    update_status_column(wb, idx, 42, "")
                    logger.debug(f"Sem linha para processar na linha {idx}")

            # Verifica se celular foi processado ou se não há celular
            if cel_modelo:
                # Format phone name to include brand and IMEI
                phone_name = cel_modelo
                if cel_marca and cel_imei and str(cel_imei).strip():
                    phone_name = f"Celular {str(cel_marca).strip()} - {str(cel_imei).strip()}"
                elif cel_imei and str(cel_imei).strip():
                    phone_name = f"Celular - {str(cel_imei).strip()}"
                
                # Prepara os dados do telefone
                phone_data = {
                    "name": phone_name,
                    "entities_id": entidade_final_id,
                    "users_id": user_id if user_id else 0,  # Usa 0 como fallback
                    "phonetypes_id": cel_type,
                    "states_id": cel_status  # Corrigido: states_id ao invés de status
                }

                # Busca ou cria o modelo
                if cel_modelo:
                    model_id = get_or_create_model(session, cel_modelo, "Phone")
                    if model_id:
                        phone_data["phonemodels_id"] = model_id
                    else:
                        print(c(f"⚠️ [AVISO] Não foi possível criar/encontrar o modelo '{cel_modelo}'", 'yellow'))

                # Busca ou cria o fabricante
                if cel_marca and str(cel_marca).strip():
                    manufacturer_id = get_or_create_manufacturer(session, str(cel_marca).strip())
                    if manufacturer_id:
                        phone_data["manufacturers_id"] = manufacturer_id
                    else:
                        print(c(f"⚠️ [AVISO] Não foi possível criar/encontrar o fabricante '{cel_marca}'", 'yellow'))

                # Adiciona IMEI como número serial
                if cel_imei and str(cel_imei).strip():
                    phone_data["serial"] = str(cel_imei).strip()

                if cel_coment and str(cel_coment).strip():
                    phone_data["comment"] = str(cel_coment).strip()

                phone_id, phone_error = create_asset(session, "Phone", phone_data)
                
                # Atualiza status do celular na planilha
                if phone_id:
                    update_status_column(wb, idx, 43, "OK")  # Coluna 43 = Input Mobile
                    logger.item_created("Phone", phone_id, phone_name)
                else:
                    error_msg = phone_error if phone_error else "Falha ao criar celular"
                    update_status_column(wb, idx, 43, "ERRO", error_msg)
                    logger.item_failed("Phone", error_msg)
            else:
                # Se não há celular para processar, marca como vazio
                if not cel_modelo or str(cel_modelo).strip() == "":
                    update_status_column(wb, idx, 43, "")
                    logger.debug(f"Sem celular para processar na linha {idx}")

            # Verifica se notebook foi processado ou se não há notebook  
            if nb_modelo:
                # Format computer name to include manufacturer and serial number
                computer_name = f"Notebook {str(nb_marca).strip()} - {str(nb_serial).strip()}"
                
                # Busca ou cria o modelo
                model_id = get_or_create_model(session, nb_modelo, "Computer")
                if not model_id:
                    print(c(f"⚠️ [AVISO] Não foi possível criar/encontrar o modelo '{nb_modelo}'", 'yellow'))
                    continue

                # Busca ou cria o fabricante
                manufacturer_id = get_or_create_manufacturer(session, str(nb_marca).strip())
                if not manufacturer_id:
                    print(c(f"⚠️ [AVISO] Não foi possível criar/encontrar o fabricante '{nb_marca}'", 'yellow'))
                    continue

                computer_data = {
                "name": computer_name,
                "entities_id": entidade_final_id,
                "users_id": user_id if user_id else 0,  # Usa 0 como fallback
                "is_dynamic": 0,  # Garantir que não é um computador dinâmico
                "computermodels_id": model_id,
                "manufacturers_id": manufacturer_id,
                "serial": nb_serial,
                "otherserial": nb_ativo,
                "computertypes_id": nb_type,
                "states_id": nb_status  # Corrigido: states_id ao invés de status
                }

                if nb_coment and str(nb_coment).strip():
                    computer_data["comment"] = str(nb_coment).strip()

                # Cria o computador
                computer_id, computer_error = create_asset(session, "Computer", computer_data)
                
                # Atualiza status do notebook na planilha
                if computer_id:
                    update_status_column(wb, idx, 44, "OK")  # Coluna 44 = Input Notebook
                    logger.item_created("Computer", computer_id, computer_name)
                else:
                    error_msg = computer_error if computer_error else "Falha ao criar notebook"
                    update_status_column(wb, idx, 44, "ERRO", error_msg)
                    logger.item_failed("Computer", error_msg)
                
                # Se o computador foi criado com sucesso
                if computer_id:
                    # Processa contrato do notebook
                    if contrato_notebook and str(contrato_notebook).strip():
                        supplier_id = None
                        
                        # Primeiro cria/busca o fornecedor se fornecido (na entidade raiz)
                        if fornecedor_notebook and str(fornecedor_notebook).strip():
                            supplier_id = get_or_create_supplier(session, str(fornecedor_notebook).strip())
                        
                        # Cria/busca o contrato com o fornecedor (na entidade raiz)
                        contract_id = get_or_create_contract(session, str(contrato_notebook).strip(), supplier_id=supplier_id)
                        if contract_id:
                            link_contract_to_asset(session, "Computer", computer_id, contract_id)
                    
                    # Processa informações de Management
                    if comprado_em_notebook and str(comprado_em_notebook).strip():
                        # Pegar o supplier_id do contrato do notebook
                        notebook_supplier_id = None
                        if contrato_notebook and str(contrato_notebook).strip():
                            # Buscar o supplier do contrato do notebook
                            if fornecedor_notebook and str(fornecedor_notebook).strip():
                                notebook_supplier_id = get_or_create_supplier(session, str(fornecedor_notebook).strip())
                        
                        create_management_info(
                            session,
                            "Computer", 
                            computer_id,
                            buy_date=comprado_em_notebook,
                            supplier_id=notebook_supplier_id
                        )

                    # Linka componentes ao computador
                    link_component(computer_id, nb_armazenamento, nb_processador, nb_memoria, session)

                logger.success("Notebook e componentes processados com sucesso")
            else:
                # Se não há notebook para processar, marca como vazio
                if not nb_modelo or str(nb_modelo).strip() == "":
                    update_status_column(wb, idx, 44, "")
                    logger.debug(f"Sem notebook para processar na linha {idx}")

            logger.success(f"Linha {idx} processada com sucesso")
            total_processado += 1
            total_sucesso += 1
            
            # Salva as alterações de status na planilha após cada linha
            save_excel_file(wb)
        except Exception as e:
            logger.error(f"Erro na linha {idx}: {e}")
            total_processado += 1
            total_erro += 1
            
            # Em caso de erro geral, marca todas as colunas com a descrição do erro
            error_description = f"Erro geral: {str(e)[:50]}..."  # Limita tamanho da mensagem
            update_status_column(wb, idx, 41, "ERRO", error_description)  # Input User
            update_status_column(wb, idx, 42, "ERRO", error_description)  # Input Line  
            update_status_column(wb, idx, 43, "ERRO", error_description)  # Input Mobile
            update_status_column(wb, idx, 44, "ERRO", error_description)  # Input Notebook
            save_excel_file(wb)

    kill_session(session)
    
    # Usar o logger para estatísticas finais  
    logger = get_logger()
    
    # Salvamento final da planilha
    if save_excel_file(wb):
        logger.success("Planilha salva com status atualizados!")
    else:
        logger.error("Erro ao salvar planilha final!")
    
    logger.statistics(total_processado, total_sucesso, total_erro)
    logger.process_end("Processo de Input de dados no GLPI")
    
    return total_processado, total_sucesso, total_erro


if __name__ == "__main__":
    try:
        # Inicializar logger
        logger = get_logger()
        logger.process_start("Reset do GLPI")
        reset_glpi()
        
        total, sucessos, erros = main()
        
        if total > 0:
            taxa_sucesso = (sucessos / total) * 100
            logger.info(f"Taxa de sucesso: {taxa_sucesso:.1f}%")
        
        if erros > 0:
            logger.warning(f"{erros} erro(s) encontrado(s)")
            
        # Fechar o logger
        close_logger()
            
    except Exception as e:
        logger = get_logger()
        logger.critical(f"Erro fatal: {e}")
        close_logger()