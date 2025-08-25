"""
Script de automação para input de usuários, ativos, contratos e custos no GLPI via API REST.
Utiliza uma planilha XLSX como fonte de dados e realiza operações de criação/atualização com deduplicação, vinculação de entidades, grupos e perfis.

Principais funções:
- Criação de entidades em cascata (até 4 níveis)
- Criação de grupos na entidade raiz 'GI Group'
- Criação de usuários com perfil Self-Service
- Criação e atualização de ativos vinculados a entidades e usuários
- Deduplicação e atualização condicional
- Tratamento robusto de erros e saída informativa
"""

import requests
import openpyxl
import sys
import os
from remove_data import reset_glpi

# Funções para colorir terminal (Windows suporta via colorama, mas ANSI funciona no PowerShell moderno)
def c(text, color):
    colors = {
        'green': '\033[92m',
        'red': '\033[91m',
        'yellow': '\033[93m',
        'blue': '\033[94m',
        'cyan': '\033[96m',
        'reset': '\033[0m'
    }
    return f"{colors.get(color, '')}{text}{colors['reset']}"

###############################
# CONFIGURAÇÕES DO GLPI
###############################
GLPI_URL = "http://localhost/glpi/apirest.php"  # URL da API REST do GLPI
APP_TOKEN = "xOTvbn5CqY1tfQXrHdSPmUgHSBd9y6mIZyUrt9Z8"  # Token da aplicação
USER_TOKEN = "6I5gkjwvKcVaOpQOw9TqUKaX0119owQuOmJzENHd"  # Token do usuário

# Defina aqui o ID do grupo 'User' na entidade raiz 'Gi Group' antes de executar o script
GROUP_ID = 136  # <--- Altere para o ID correto do grupo 'User'

# Como obter o ID do grupo 'User' já criado:
# 1. No GLPI, acesse Administração > Grupos.
# 2. Clique no grupo 'User' na entidade 'Gi Group'.
# 3. O ID estará na URL (exemplo: .../group.form.php?id=123) ou pode ser visto na listagem.

HEADERS = {
    "App-Token": APP_TOKEN,
    "Authorization": f"user_token {USER_TOKEN}"
}

###############################
# FUNÇÕES AUXILIARES
###############################
def init_session():
    """Inicia uma sessão na API do GLPI e retorna o token de sessão."""
    r = requests.get(f"{GLPI_URL}/initSession", headers=HEADERS)
    r.raise_for_status()
    return r.json()["session_token"]

def kill_session(session_token):
    """Encerra a sessão na API do GLPI."""
    requests.get(f"{GLPI_URL}/killSession", headers={**HEADERS, "Session-Token": session_token})

def get_or_create(session_token, endpoint, search_field, search_value, payload_extra=None):
    """
    Busca um item pelo campo especificado. Se não existir, cria o item.
    Retorna o ID do item encontrado ou criado.
    
    Args:
        session_token: Token da sessão GLPI
        endpoint: Endpoint da API (Entity, Group, User, etc)
        search_field: Campo a ser usado na busca
        search_value: Valor a ser buscado
        payload_extra: Dados adicionais para criação/busca (ex: entities_id)
    
    Returns:
        int: ID do item encontrado ou criado
        None: Se não foi possível encontrar ou criar o item
    """
    headers = {**HEADERS, "Session-Token": session_token}
    print(c(f"🔎 [BUSCA] Procurando {endpoint} '{search_value}'...", 'cyan'))
    
    # Para grupos, tenta busca direta primeiro
    if endpoint == "Group":
        try:
            groups = requests.get(f"{GLPI_URL}/Group", headers=headers).json()
            if isinstance(groups, list):
                print(c(f"[DEBUG] Verificando na lista de {len(groups)} grupos...", 'yellow'))
                for group in groups:
                    if group.get("name") == search_value:
                        entity_id = group.get("entities_id")
                        if not payload_extra or not payload_extra.get("entities_id") or str(entity_id) == str(payload_extra.get("entities_id")):
                            found_id = int(group.get("id"))
                            print(c(f"✅ [OK] {endpoint} '{search_value}' encontrado (ID: {found_id})", 'green'))
                            return found_id
        except Exception as e:
            print(c(f"[DEBUG] Erro na busca direta de grupos: {e}", 'yellow'))

    # Busca via API de busca
    params = {"criteria[0][field]": 1, "criteria[0][searchtype]": "equals", "criteria[0][value]": search_value}
    # Para entidades/grupos, busca também por entities_id se fornecido
    if endpoint in ["Entity", "Group"] and payload_extra and "entities_id" in payload_extra:
        params["criteria[1][field]"] = 4
        params["criteria[1][searchtype]"] = "equals"
        params["criteria[1][value]"] = payload_extra["entities_id"]
    search = requests.get(f"{GLPI_URL}/search/{endpoint}", headers=headers, params=params)
    resp = search.json()
    # Corrige caso a resposta seja uma lista inesperada
    if isinstance(resp, list):
        print(c(f"⚠️ [AVISO] Resposta inesperada da API (lista) ao buscar {endpoint} '{search_value}'. Buscando manualmente...", 'yellow'))
        for item in resp:
            if isinstance(item, dict):
                if (str(item.get("1", "")) == str(search_value)) or (item.get("name", "") == str(search_value)):
                    item_id = int(item.get("id", item.get("2", 0)))
                    print(c(f"✅ [OK] {endpoint} '{search_value}' encontrado (ID: {item_id})", 'green'))
                    return item_id
        resp = {"totalcount": 0}
    if resp.get("totalcount", 0) > 0:
        item_id = int(resp["data"][0].get("id", resp["data"][0].get("2", 0)))
        print(c(f"✅ [OK] {endpoint} '{search_value}' encontrado (ID: {item_id})", 'green'))
        return item_id
    # Para entidades, verifica se já existe via busca direta primeiro
    if endpoint == "Entity":
        try:
            entities_list = requests.get(f"{GLPI_URL}/Entity", headers=headers).json()
            if isinstance(entities_list, list):
                print(c(f"[DEBUG] Verificando na lista de {len(entities_list)} entidades antes de criar...", 'yellow'))
                for entity in entities_list:
                    if entity.get("name") == search_value:
                        parent_id = entity.get("entities_id", "") or entity.get("parent_id", "")
                        if not payload_extra or not payload_extra.get("entities_id") or str(parent_id) == str(payload_extra.get("entities_id")):
                            found_id = int(entity.get("id"))
                            print(c(f"✅ [OK] {endpoint} '{search_value}' já existe (ID: {found_id})", 'green'))
                            return found_id
        except Exception as e:
            print(c(f"[DEBUG] Erro na verificação prévia: {e}", 'yellow'))

    print(c(f"🆕 [CRIANDO] {endpoint} '{search_value}'...", 'blue'))
    payload = {"input": {search_field: search_value}}
    if payload_extra:
        payload["input"].update(payload_extra)
    r = requests.post(f"{GLPI_URL}/{endpoint}", headers=headers, json=payload)
    try:
        r.raise_for_status()
        item_id = r.json()["id"]
        print(c(f"✅ [OK] {endpoint} '{search_value}' criado (ID: {item_id})", 'green'))
        return item_id
    except Exception as e:
        print(c(f"❌ [ERRO] Não foi possível criar {endpoint} '{search_value}'. Resposta da API:", 'red'))
        print(r.text)
        # Se erro for de duplicidade, buscar novamente e retornar o ID
        if "já existe" in r.text or "already exists" in r.text or "Duplicate entry" in r.text:
            print(c(f"🔄 [BUSCA EXTRA] {endpoint} '{search_value}' já existe, buscando ID...", 'yellow'))
            # Busca manual considerando nome e parent (entities_id)
            search_params = {"criteria[0][field]": 1, "criteria[0][searchtype]": "equals", "criteria[0][value]": search_value}
            if endpoint in ["Entity", "Group"] and payload_extra and "entities_id" in payload_extra:
                search_params["criteria[1][field]"] = 4
                search_params["criteria[1][searchtype]"] = "equals"
                search_params["criteria[1][value]"] = payload_extra["entities_id"]
            search2 = requests.get(f"{GLPI_URL}/search/{endpoint}", headers=headers, params=search_params)
            print(c(f"[DEBUG] Resposta da busca por nome+parent: {search2.text}", 'yellow'))
            resp2 = search2.json()
            found_id = None
            if isinstance(resp2, list):
                for item in resp2:
                    if isinstance(item, dict):
                        nome_match = (str(item.get("1", "")) == str(search_value)) or (item.get("name", "") == str(search_value))
                        if not nome_match and "completename" in item:
                            nome_match = item["completename"].endswith(str(search_value))
                        if nome_match:
                            parent_id = item.get("entities_id") or item.get("4") or item.get("parent_id")
                            parent_match = True
                            if endpoint in ["Entity", "Group"] and payload_extra and "entities_id" in payload_extra:
                                parent_match = str(parent_id) == str(payload_extra["entities_id"])
                            if parent_match:
                                found_id = int(item.get("id", item.get("2", 0)))
                                break
                if found_id:
                    print(c(f"✅ [OK] {endpoint} '{search_value}' encontrado após erro (ID: {found_id})", 'green'))
                    return found_id
                print(c(f"❌ [ERRO] {endpoint} '{search_value}' não encontrado após erro de duplicidade.", 'red'))
                print(c(f"[DEBUG] Resposta da busca extra: {resp2}", 'yellow'))
                # Fallback: buscar todas entidades e filtrar manualmente por parent e nome
                if endpoint == "Entity" and payload_extra and "entities_id" in payload_extra:
                    print(c(f"[DEBUG] Fallback: buscando todas entidades e filtrando por parent ID {payload_extra['entities_id']} e nome '{search_value}'...", 'yellow'))
                    headers = {**HEADERS, "Session-Token": session_token}
                    resp_fallback = requests.get(f"{GLPI_URL}/search/Entity", headers=headers).json()
                    if resp_fallback.get("totalcount", 0) > 0:
                        for data in resp_fallback["data"]:
                            nome_match = (str(data.get("1", "")) == str(search_value)) or (data.get("name", "") == str(search_value))
                            parent_id = data.get("entities_id") or data.get("parent_id") or data.get("4")
                            parent_match = str(parent_id) == str(payload_extra["entities_id"])
                            if not nome_match and "completename" in data:
                                nome_match = data["completename"].endswith(str(search_value))
                            if nome_match and parent_match:
                                fallback_id = int(data.get("id", data.get("2", 0)))
                                print(c(f"✅ [OK] {endpoint} '{search_value}' encontrado via fallback (ID: {fallback_id})", 'green'))
                                return fallback_id
                    print(c(f"❌ [ERRO] {endpoint} '{search_value}' não encontrado via fallback.", 'red'))
            elif resp2.get("totalcount", 0) > 0:
                for data in resp2["data"]:
                    nome_match = (str(data.get("1", "")) == str(search_value)) or (data.get("name", "") == str(search_value))
                    if not nome_match and "completename" in data:
                        nome_match = data["completename"].endswith(str(search_value))
                    if nome_match:
                        parent_id = data.get("entities_id") or data.get("4") or data.get("parent_id")
                        parent_match = True
                        if endpoint in ["Entity", "Group"] and payload_extra and "entities_id" in payload_extra:
                            parent_match = str(parent_id) == str(payload_extra["entities_id"])
                        if parent_match:
                            found_id = int(data.get("id", data.get("2", 0)))
                            break
                if found_id:
                    print(c(f"✅ [OK] {endpoint} '{search_value}' encontrado após erro (ID: {found_id})", 'green'))
                    return found_id
                print(c(f"❌ [ERRO] {endpoint} '{search_value}' não encontrado após erro de duplicidade.", 'red'))
                print(c(f"[DEBUG] Resposta da busca extra: {resp2}", 'yellow'))
            # Busca direta na API
            print(c(f"[DEBUG] Tentando busca direta via API...", 'yellow'))
            if endpoint == "Entity":
                try:
                    all_entities = requests.get(f"{GLPI_URL}/Entity", headers=headers).json()
                    if isinstance(all_entities, list):
                        print(c(f"[DEBUG] Encontradas {len(all_entities)} entidades, procurando '{search_value}'...", 'yellow'))
                        # Primeiro procura por correspondência exata de nome e parent
                        for entity in all_entities:
                            nome = entity.get("name", "")
                            eid = entity.get("id", "")
                            parent = entity.get("entities_id", "") or entity.get("parent_id", "") or entity.get("4", "")
                            if nome == search_value:
                                if not payload_extra or not payload_extra.get("entities_id") or str(parent) == str(payload_extra.get("entities_id", "")):
                                    print(c(f"✅ [OK] Entity '{search_value}' encontrado via API direta (ID: {eid})", 'green'))
                                    return int(eid)
                        # Se não encontrou, imprime todas para debug
                        print(c(f"[DEBUG] Listando todas entidades:", 'yellow'))
                        for entity in all_entities:
                            nome = entity.get("name", "")
                            eid = entity.get("id", "")
                            parent = entity.get("entities_id", "") or entity.get("parent_id", "") or entity.get("4", "")
                            print(c(f"  - Nome: {nome} | ID: {eid} | Parent: {parent}", 'yellow'))
                except Exception as e:
                    print(c(f"[DEBUG] Erro na busca direta: {e}", 'yellow'))

            # Se busca direta falhar, tenta busca alternativa
            print(c(f"[DEBUG] Tentando busca alternativa com diferentes campos...", 'yellow'))
            for field in ["name", "1", "2", "completename"]:
                alt_search_params = {"criteria[0][field]": field, "criteria[0][searchtype]": "equals", "criteria[0][value]": search_value}
                if endpoint == "Entity" and payload_extra and "entities_id" in payload_extra:
                    alt_search_params.update({
                        "criteria[1][link]": "AND",
                        "criteria[1][field]": "entities_id",
                        "criteria[1][searchtype]": "equals",
                        "criteria[1][value]": payload_extra["entities_id"]
                    })
                alt_search_response = requests.get(f"{GLPI_URL}/search/{endpoint}", headers=headers, params=alt_search_params)
                print(c(f"[DEBUG] Resposta da busca alternativa (campo {field}): {alt_search_response.text}", 'yellow'))
            try:
                alt_resp = alt_search_response.json()
                if isinstance(alt_resp, list) and len(alt_resp) > 0:
                    for item in alt_resp:
                        if isinstance(item, dict):
                            nome = item.get("name", item.get("1", ""))
                            eid = item.get("id", item.get("2", ""))
                            parent = item.get("entities_id", "") or item.get("parent_id", "") or item.get("4", "")
                            if nome == search_value and str(parent) == str(payload_extra.get("entities_id", "")):
                                print(c(f"✅ [OK] {endpoint} '{search_value}' encontrado via busca alternativa (ID: {eid})", 'green'))
                                return int(eid)
                elif alt_resp.get("totalcount", 0) > 0:
                    for data in alt_resp["data"]:
                        nome = data.get("name", data.get("1", ""))
                        eid = data.get("2", data.get("id", ""))
                        parent = data.get("entities_id", "") or data.get("parent_id", "") or data.get("4", "")
                        if nome == search_value and str(parent) == str(payload_extra.get("entities_id", "")):
                            print(c(f"✅ [OK] {endpoint} '{search_value}' encontrado via busca alternativa (ID: {eid})", 'green'))
                            return int(eid)
                elif alt_resp.get("totalcount", 0) > 0:
                    print(c(f"[DEBUG] Itens retornados na busca alternativa:", 'yellow'))
                    for data in alt_resp["data"]:
                        nome = data.get("name", data.get("1", ""))
                        eid = data.get("id", data.get("2", ""))
                        parent = data.get("entities_id", data.get("parent_id", ""))
                        print(c(f"  - Nome: {nome} | ID: {eid} | Parent: {parent}", 'yellow'))
                        nome_match = (str(data.get("1", "")) == str(search_value)) or (nome == str(search_value))
                        if not nome_match and "completename" in data:
                            nome_match = data["completename"].endswith(str(search_value))
                        if nome_match:
                            alt_found_id = int(eid)
                    if alt_found_id:
                        print(c(f"✅ [OK] {endpoint} '{search_value}' encontrado (ID: {alt_found_id}) [busca alternativa]", 'green'))
                        return alt_found_id
            except Exception as e:
                print(c(f"[DEBUG] Erro ao processar resposta JSON alternativa: {e}", 'yellow'))
            print(c(f"❌ [ERRO] {endpoint} '{search_value}' não encontrado após erro de duplicidade.", 'red'))
            return None

def create_entity_hierarchy(session_token, entidade_a, entidade_b=None, entidade_c=None, entidade_d=None):
    """
    Cria entidades em cascata (até 4 níveis) e retorna o ID da entidade mais profunda criada.
    """
    print(c("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", 'yellow'))
    print(c("🏢 [ETAPA] Criando hierarquia de entidades...", 'yellow'))
    eid_a = get_or_create(session_token, "Entity", "name", entidade_a)
    if eid_a is None:
        print(c(f"❌ [ERRO] Não foi possível criar ou encontrar a entidade A '{entidade_a}'. Hierarquia abortada.", 'red'))
        return None
    # Busca entidade B antes de criar, para evitar duplicidade
    eid_b = None
    if entidade_b:
        print(c(f"🔎 [BUSCA] Procurando/reativando Entity B '{entidade_b}' sob '{entidade_a}'...", 'cyan'))
        eid_b = get_or_create(session_token, "Entity", "name", entidade_b, {"entities_id": eid_a})
        if eid_b:
            print(c(f"♻️ [INFO] Entity B '{entidade_b}' será reutilizada para entidades C filhas.", 'yellow'))
        else:
            print(c(f"❌ [ERRO] Não foi possível criar ou encontrar a entidade B '{entidade_b}'. Hierarquia abortada.", 'red'))
            return eid_a
    # Cria entidade C como filha de B
    eid_c = None
    if entidade_c:
        print(c(f"🔎 [BUSCA] Procurando ou criando Entity C '{entidade_c}' sob Entity B '{entidade_b}' (ID: {eid_b})...", 'cyan'))
        # Busca manual por entidade C sob o parent correto
        headers = {**HEADERS, "Session-Token": session_token}
        params_c = {"criteria[0][field]": 1, "criteria[0][searchtype]": "equals", "criteria[0][value]": entidade_c,
                    "criteria[1][field]": 4, "criteria[1][searchtype]": "equals", "criteria[1][value]": eid_b}
        search_c = requests.get(f"{GLPI_URL}/search/Entity", headers=headers, params=params_c)
        resp_c = search_c.json()
        if resp_c.get("totalcount", 0) > 0:
            eid_c = int(resp_c["data"][0].get("id", resp_c["data"][0].get("2", 0)))
            print(c(f"♻️ [INFO] Entity C '{entidade_c}' já existe sob Entity B '{entidade_b}' (ID: {eid_c}). Reutilizando.", 'yellow'))
        else:
            eid_c = get_or_create(session_token, "Entity", "name", entidade_c, {"entities_id": eid_b})
            if eid_c:
                print(c(f"✅ [OK] Entity C '{entidade_c}' criada sob Entity B '{entidade_b}' (ID: {eid_b}).", 'green'))
            else:
                print(c(f"❌ [ERRO] Não foi possível criar ou encontrar a entidade C '{entidade_c}'. Hierarquia abortada.", 'red'))
                return eid_b if eid_b is not None else eid_a
    eid_d = None
    if entidade_d:
        print(c(f"🔎 [BUSCA] Criando Entity D '{entidade_d}' sob '{entidade_c}'...", 'cyan'))
        eid_d = get_or_create(session_token, "Entity", "name", entidade_d, {"entities_id": eid_c})
        if eid_d:
            print(c(f"✅ [OK] Entity D '{entidade_d}' criada/reutilizada sob '{entidade_c}'.", 'green'))
        else:
            print(c(f"❌ [ERRO] Não foi possível criar ou encontrar a entidade D '{entidade_d}'. Hierarquia abortada.", 'red'))
            return eid_c if eid_c is not None else (eid_b if eid_b is not None else eid_a)
    # Retorna o ID da entidade mais profunda criada
    return eid_d or eid_c or eid_b or eid_a

def create_user(session_token, name, group, profile_id, entity_id):
    """
    Cria usuário e configura suas permissões.
    
    Args:
        session_token: Token da sessão GLPI
        name: Nome do usuário
        group: Nome do grupo (será usado o GROUP_ID global)
        profile_id: ID do perfil a ser associado
        entity_id: ID da entidade onde o usuário será criado
    
    Returns:
        int: ID do usuário criado/encontrado
        None: Se houve erro na criação
    """
    print(c("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", 'yellow'))
    print(c(f"👤 [ETAPA] Criando usuário '{name}'...", 'yellow'))
    
    # Validação inicial do profile_id
    if not profile_id:
        print(c(f"⚠️ [AVISO] Profile ID não fornecido para '{name}'. Usando perfil padrão.", 'yellow'))
        profile_id = 1  # Perfil Self-Service como fallback
        
    # Cria ou busca o usuário primeiro
    user_id = get_or_create(session_token, "User", "name", name, {"entities_id": entity_id})

    # Usa o ID do grupo definido na variável GROUP_ID
    group_id = GROUP_ID
    if group_id:
        print(c(f"🔗 [INFO] Grupo padrão para usuário '{name}' (ID do grupo definido: {group_id})", 'cyan'))
    else:
        print(c(f"❌ [ERRO] GROUP_ID não definido. Usuário '{name}' não será vinculado ao grupo.", 'red'))

    # headers para requisições
    headers = {**HEADERS, "Session-Token": session_token}

    # Busca o ID do perfil Self-Service
    print(c(f"🛡️ [INFO] Perfil para usuário '{name}': {profile_id}", 'cyan'))
    perfil_ok = True
    if profile_id:
        perfil_search = requests.get(f"{GLPI_URL}/search/Profile", headers=headers,
            params={"criteria[0][field]": 1, "criteria[0][searchtype]": "equals", "criteria[0][value]": profile_id})
        perfil_resp = perfil_search.json()
        if isinstance(perfil_resp, list):
            perfil_resp = {"totalcount": 0}
        if perfil_resp.get("totalcount", 0) == 0:
            print(c(f"❌ [ERRO] Perfil ID {profile_id} não existe no GLPI!", 'red'))
            perfil_ok = False

    # 1. Vincula o usuário ao grupo via Group_User
    if group_id:
        group_user_payload = {"input": {"users_id": user_id, "groups_id": group_id}}
        r_group_user = requests.post(f"{GLPI_URL}/Group_User", headers=headers, json=group_user_payload)
        if r_group_user.status_code in [200, 201]:
            print(c(f"✅ [OK] Usuário '{name}' vinculado ao grupo secundário!", 'green'))
        elif r_group_user.status_code == 400 and ("already exists" in r_group_user.text or "Duplicate entry" in r_group_user.text):
            print(c(f"ℹ️ [INFO] Associação secundária já existe para '{name}'.", 'blue'))
        else:
            print(c(f"❌ [ERRO] Falha ao vincular grupo secundário para '{name}': {r_group_user.text}", 'red'))
    # 2. Atualiza o campo groups_id do usuário
    update_payload = {"input": {}}
    if group_id:
        update_payload["input"]["groups_id"] = group_id  # Grupo padrão
    if perfil_ok and profile_id:
        update_payload["input"]["profiles_id"] = profile_id
    # Sempre atualiza a entidade do usuário para a entidade mais profunda
    update_payload["input"]["entities_id"] = entity_id
    if update_payload["input"]:
        r_update = requests.put(f"{GLPI_URL}/User/{user_id}", headers=headers, json=update_payload)
        if r_update.status_code == 200:
            print(c(f"✅ [OK] Usuário '{name}' atualizado com grupo/perfil padrão!", 'green'))
        else:
            print(c(f"❌ [ERRO] Falha ao atualizar grupo/perfil padrão do usuário '{name}': {r_update.text}", 'red'))

    # Vincula o perfil (autorização) à entidade correta
    if perfil_ok and profile_id:
        profile_payload = {
            "input": {
                "profiles_id": profile_id,
                "entities_id": entity_id
            }
        }
        r_profile = requests.post(f"{GLPI_URL}/User/{user_id}/Profile", headers=headers, json=profile_payload)
        if r_profile.status_code == 201:
            print(c(f"✅ [OK] Autorização do perfil vinculada à entidade correta para '{name}'!", 'green'))
        elif r_profile.status_code == 400 and ("already exists" in r_profile.text or "Duplicate entry" in r_profile.text):
            print(c(f"ℹ️ [INFO] Autorização do perfil já existe para '{name}' na entidade correta.", 'blue'))
        else:
            print(c(f"❌ [ERRO] Falha ao vincular autorização do perfil para '{name}': {r_profile.text}", 'red'))
    return user_id

def create_asset(session_token, asset_type, payload):
    """
    Cria ou atualiza um ativo (Line, Phone, Computer) vinculado à entidade e usuário.
    Retorna o ID do ativo criado ou atualizado.
    """
    headers = {**HEADERS, "Session-Token": session_token}
    print(c("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", 'yellow'))
    print(c(f"💻 [ETAPA] Criando ativo do tipo {asset_type}...", 'yellow'))
    search_value = payload.get("name")
    search = requests.get(f"{GLPI_URL}/search/{asset_type}", headers=headers,
                         params={"criteria[0][field]": 1,
                                 "criteria[0][searchtype]": "equals",
                                 "criteria[0][value]": search_value})
    resp = search.json()
    if resp.get("totalcount", 0) > 0:
        asset_id = int(resp["data"][0].get("id", resp["data"][0].get("2", 0)))
        print(c(f"♻️ [OK] {asset_type} '{search_value}' já existe (ID: {asset_id}), atualizando...", 'green'))
        requests.put(f"{GLPI_URL}/{asset_type}/{asset_id}", headers=headers, json={"input": payload})
        return asset_id
    print(c(f"🆕 [CRIANDO] {asset_type} '{search_value}'...", 'blue'))
    r = requests.post(f"{GLPI_URL}/{asset_type}", headers=headers, json={"input": payload})
    try:
        r.raise_for_status()
        asset_id = r.json()["id"]
        print(c(f"✅ [OK] {asset_type} '{search_value}' criado (ID: {asset_id})", 'green'))
        return asset_id
    except Exception as e:
        print(c(f"❌ [ERRO] Não foi possível criar {asset_type} '{search_value}'. Resposta da API:", 'red'))
        print(r.text)
        raise

###############################
# LEITURA DA PLANILHA E PROCESSAMENTO
###############################
def process_excel(file_path="input_glpi.xlsx"):
    """
    Processa a planilha XLSX e realiza input dos dados no GLPI via API.
    
    Args:
        file_path: Caminho para o arquivo XLSX com os dados
        
    A planilha deve conter as seguintes colunas:
    - Nome do usuário
    - Entidade A (obrigatória)
    - Entidade B (opcional)
    - Entidade C (opcional)
    - Linha telefônica
    - Celular modelo
    - Celular IMEI
    - Notebook modelo
    - Notebook marca
    - Notebook serial
    """
    print(c("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", 'yellow'))
    print(c("🚀 [INICIANDO] Processamento da planilha XLSX...", 'yellow'))
    
    session = None
    try:
        if not os.path.exists(file_path):
            print(c(f"❌ [ERRO] Arquivo '{file_path}' não encontrado!", 'red'))
            return
            
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        if sheet.max_row < 2:
            print(c("❌ [ERRO] Planilha vazia ou sem dados!", 'red'))
            return
        session = init_session()
    except Exception as e:
        print(c(f"❌ [ERRO] Falha ao processar arquivo: {str(e)}", 'red'))
        if session:
            kill_session(session)
        return

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        print(c(f"\n📄 [LINHA {idx}] Processando dados...", 'blue'))
        try:
            # Desempacota os campos da linha
            nome, ent_a, ent_b, ent_c, linha, cel_modelo, cel_imei, nb_modelo, nb_marca, nb_serial = row

            # Validação de campos obrigatórios
            if not ent_a:
                print(c(f"❌ [ERRO] Entidade A obrigatória na linha {idx}. Pulando linha.", 'red'))
                continue

            # Cria entidades em cascata (até 3 níveis)
            entidade_id = create_entity_hierarchy(session, ent_a, ent_b, ent_c)

            # Cria usuário e vincula sempre ao grupo 'User' e ao perfil Self-Service
            user_id = None
            if nome:
                perfil_id = 1  # ID do perfil Self-Service
                user_id = create_user(session, nome, 'User', perfil_id, entidade_id)

            # Cria ativos vinculados à entidade/usuário (apenas se campo preenchido)
            if linha:
                create_asset(session, "Line", {"name": linha, "entities_id": entidade_id, "users_id": user_id if user_id else None})

            if cel_modelo and cel_imei:
                create_asset(session, "Phone", {"name": cel_modelo, "imei": cel_imei, "entities_id": entidade_id, "users_id": user_id if user_id else None})

            if nb_modelo and nb_serial:
                create_asset(session, "Computer", {"name": nb_modelo, "serial": nb_serial, "manufacturers_id": nb_marca,
                                                   "entities_id": entidade_id, "users_id": user_id if user_id else None})

            print(c(f"✅ [OK] Linha {idx} processada!", 'green'))
        except Exception as e:
            print(c(f"❌ [ERRO] Falha ao processar linha {idx}: {e}", 'red'))

    kill_session(session)
    print(c("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", 'yellow'))
    print(c("🏁 [FINALIZADO] Sessão encerrada e processamento concluído! 🚀", 'yellow'))

###############################
# EXECUÇÃO PRINCIPAL
###############################
if __name__ == "__main__":
    reset_glpi()
    process_excel("input_glpi.xlsx")
    print(c("\n✅ Processamento concluído com sucesso! 🚀", 'green'))
