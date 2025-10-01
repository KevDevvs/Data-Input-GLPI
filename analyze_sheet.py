"""
Script para analisar toda a estrutura da planilha
"""
import openpyxl
from helper.colors import c

def analyze_spreadsheet():
    """Analisa toda a estrutura da planilha"""
    print(c("🔍 Analisando estrutura completa da planilha...", 'yellow'))
    
    try:
        # Carrega a planilha
        wb = openpyxl.load_workbook("input_glpi.xlsx")
        sheet = wb.active
        
        # Pega os cabeçalhos da primeira linha
        headers = []
        first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        print(c("\n📊 Estrutura da planilha (Cabeçalhos):", 'cyan'))
        print(c("Coluna | Nome do Campo", 'white'))
        print(c("-" * 50, 'white'))
        
        for idx, header in enumerate(first_row, start=1):
            if header:  # Só mostra colunas que têm cabeçalho
                print(f"{idx:6d} | {header}")
        
        print(c(f"\nTotal de colunas com dados: {len([h for h in first_row if h])}", 'green'))
        
        # Mostra alguns exemplos de dados
        print(c("\n📝 Exemplos de dados (linha 2):", 'cyan'))
        if sheet.max_row >= 2:
            second_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            for idx, (header, value) in enumerate(zip(first_row, second_row), start=1):
                if header and value:  # Só mostra campos que têm cabeçalho e valor
                    print(f"{idx:6d} | {header:30} | {value}")
                
    except Exception as e:
        print(c(f"❌ Erro ao analisar planilha: {e}", 'red'))

if __name__ == "__main__":
    analyze_spreadsheet()